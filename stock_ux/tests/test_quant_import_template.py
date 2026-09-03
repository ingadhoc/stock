from odoo.tests import tagged
from odoo.tests.common import TransactionCase
from odoo.tools.misc import file_path
from openpyxl import load_workbook


@tagged("stock_ux_quant_import")
class TestQuantImportTemplate(TransactionCase):
    """The offered template has to be importable as it is downloaded."""

    # what each column of the template is meant to fill in
    COLUMNS = {
        "id": None,
        "Producto": "product_id",
        "Ubicación": "location_id",
        "Número de serie/lote": "lot_id",
        "Contado": "inventory_quantity",
    }

    def test_template_columns_can_all_be_imported(self):
        quants = self.env["stock.quant"]
        template = quants.get_import_templates()[0]["template"]
        self.assertEqual(template, "/stock_ux/static/xls/stock_quant.xlsx")

        sheet = load_workbook(file_path(template.lstrip("/"))).worksheets[0]
        headers = [cell.value for cell in next(sheet.rows)]
        self.assertEqual(headers, list(self.COLUMNS), "the template columns changed")
        self.assertEqual(sheet.max_row, 1, "an example row would be imported as a count")

        importable = quants._get_inventory_fields_create()
        for header, field in self.COLUMNS.items():
            if field:
                self.assertIn(
                    field,
                    importable,
                    "a column an inventory adjustment cannot take fails the whole file",
                )
